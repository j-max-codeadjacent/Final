from __future__ import division
import aa_authorwords as aa
import nltk
from openpyxl import Workbook
from openpyxl.compat import range

semicolon_avg = {}
for key in aa.title_keys:
	sent_token = nltk.sent_tokenize(aa.lc_authorwords[key])
	num_of_sent = len(sent_token)
	semi_count = 0
	for sent in sent_token:
		if ';' in sent:
			semi_count += 1
	semi_avg = semi_count/num_of_sent
	semicolon_avg[key] = semi_avg




wb = Workbook()
ws1 = wb.active
ws1.cell(row=1, column=1).value = 'title'
ws1.cell(row=1, column=2).value = 'semi/sent'

entry = 2
while entry <=(len(aa.title_keys)+1):
	for key in aa.title_keys:
		ws1.cell(row = entry, column = 1).value = key
		ws1.cell(row = entry, column = 2).value = semicolon_avg[key]
		entry +=1

wb.save('exceldata/semicolon.xlsx')
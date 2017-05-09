import openpyxl as op

titles = {
'middlemarch':("George Eliot", "1872"), 
'portraitofalady':("Henry James", "1881"),
'huckleberryfinn':("Mark Twain", '1884'),
'emma':('Jane Austen', '1815'),
'mysteriesofudolpho':('Ann Radcliffe', '1794'),
'uncletomscabin':('Harriet Beecher Stowe', '1852'),
'lastofthemohicans':('James Fenimore Cooper', '1826'),
'mobydick':('Herman Melville', '1851'),
'janeeyre':('Charlotte Bronte', '1847'),
'fallofthehouseofusher':('Edgar Allan Poe', '1839'),
'nature':('Ralph Waldo Emerson', '1836'),
'sartorresartus':('Thomas Carlyle', '1836'),
'walden':('Henry David Thoreau', '1854'),
'adventuresofsherlockholmes':('Arthur Conan Doyle', '1892'),
'shamela':('Henry Fielding', '1741'),
'pamela':('Samuel Richardson', '1740'),
'gulliverstravels':('Jonathan Swift', '1726'),
'callofthewild':('Jack London', '1903'),
'treasureisland':('Robert Louis Stevenson', '1883'),
'scarletletter':('Nathaniel Hawthorne', '1850'),
'olaudahequiano':('Olaudah Equiano','1789'),
'frontierinamericanhistory':('Frederick Jackson Turner', '1921'),
'lifeoffrederickdouglass':('Frederick Douglass', '1845'),
'annakarenina':('Leo Tolstoy', '1877'),
'donquixote':('Miguel de Cervantes', '1601'),
'brotherskaramazov':('Fyodor Dostoyevsky', '1880'),
'lifeofjohnson':('James Boswell', '1791'),
'taleoftwocities':('Charles Dickens', '1859'),
'narrativeofthecaptivity':('Mary Rowlandson', '1682'),
'ageofinnocence':('Edith Wharton', '1920'),
'soulsofblackfolk':('W. E. B. Du Bois', '1903'),
'ontheoriginofspecies':('Charles Darwin', '1859'),
'threecontributions':('Sigmund Freud', '1905'),
'waroftheworlds':('H. G. Wells','1898'),
'heartofdarkness':('Joseph Conrad', '1899'),
'metamorphosis':('Franz Kafka', '1915'),
'frankenstein':('Mary Shelley', '1818'),
'alicesadventures':('Lewis Carroll', '1865'),
'yellowwallpaper':('Charlotte Perkins Gilman', '1892'),
'autobiography':('John Stuart Mill', '1873'),
'anenquiry':('David Hume', '1748'),
'josephandrews':('Henry Fielding', '1742'),
'rasselas':('Samuel Johnson', '1759')
}

wb = op.Workbook()
ws = wb.create_sheet('authornames')

entry = 1
for x in titles:
	ws.cell(row = entry, column = 1).value = x
	ws.cell(row = entry, column = 2).value = titles[x][0]
	entry += 1
wb.save('authornames.xlsx')




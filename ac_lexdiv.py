from __future__ import division
import aa_authorwords as aa
import nltk
from openpyxl import Workbook
from openpyxl.compat import range

lexical_div = {}
for key in aa.title_keys:
	tokens = nltk.word_tokenize(aa.lc_authorwords[key])
	lex_div = len(set(tokens))/len(tokens)
	lexical_div[key] = lex_div




wb = Workbook()
ws1 = wb.create_sheet('lexical_diversity')
ws1.cell(row=1, column=1).value = 'title'
ws1.cell(row=1, column=2).value = 'lexical_diversity'

entry = 2
while entry <=(len(aa.title_keys)+1):
	for key in aa.title_keys:
		ws1.cell(row = entry, column = 1).value = key
		ws1.cell(row = entry, column = 2).value = lexical_div[key]
		entry +=1

wb.save('exceldata/lex_div.xlsx')






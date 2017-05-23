# -- coding: utf-8 
from __future__ import division
import pandas as pd
from sklearn.feature_extraction.stop_words import ENGLISH_STOP_WORDS
import string
import nltk
from nltk.stem.snowball import SnowballStemmer
from openpyxl import Workbook
from openpyxl.compat import range



title_keys = [
'middlemarch', 
"portraitofalady", 
"huckleberryfinn",
'emma',
'mysteriesofudolpho',
'uncletomscabin',
'lastofthemohicans',
'mobydick',
#'janeeyre', #this is not utf8
#'fallofthehouseofusher',
'nature',
'sartorresartus',
'walden',
'adventuresofsherlockholmes',
#'shamela',
'pamela',
'gulliverstravels',
'callofthewild',
'treasureisland',
'scarletletter',
'olaudahequiano',
'frontierinamericanhistory',
'lifeoffrederickdouglass',
'annakarenina',
'donquixote',
'brotherskaramazov',
'lifeofjohnson',
'taleoftwocities',
'narrativeofthecaptivity',
'ageofinnocence',
'soulsofblackfolk',
'ontheoriginofspecies',
'threecontributions',
'waroftheworlds',
'heartofdarkness',
'metamorphosis',
'frankenstein', 
'alicesadventures',
'yellowwallpaper',
'autobiography',
'anenquiry',
'josephandrews',
'rasselas'


]

titles = {
'middlemarch':("George Eliot", "1872"), 
'portraitofalady':("Henry James", "1881"),
'huckleberryfinn':("Mark Twain", '1884'),
'emma':('Jane Austen', '1815'),
'mysteriesofudolpho':('Ann Radcliffe', '1794'),
'uncletomscabin':('Harriet Beecher Stowe', '1852'),
'lastofthemohicans':('James Fenimore Cooper', '1826'),
'mobydick':('Herman Melville', '1851'),
'janeeyre':('Charlotte Brontë', '1847'),
'fallofthehouseofusher':('Edgar Allan Poe', '1839'),
'nature':('Ralph Waldo Emerson', '1836'),
'sartorresartus':('Thomas Carlyle', '1836'),
'walden':('Henry David Thoreau', '1854'),
'adventuresofsherlockholmes':('Arthur Conan Doyle', '1892'),
'shamela':('Henry Fielding', '1741'),
'pamela':('Samuel Richardson', '1740'),
'gulliverstravels':('Jonathan Swift', '1726'),
'callofthewild':('Jack London', '1903'),
'treasureisland':('Robert Louis Stevenson', '1883'),
'scarletletter':('Nathaniel Hawthorne', '1850'),
'olaudahequiano':('Olaudah Equiano','1789'),
'frontierinamericanhistory':('Frederick Jackson Turner', '1921'),
'lifeoffrederickdouglass':('Frederick Douglass', '1845'),
'annakarenina':('Leo Tolstoy', '1877'),
'donquixote':('Miguel de Cervantes', '1601'),
'brotherskaramazov':('Fyodor Dostoyevsky', '1880'),
'lifeofjohnson':('James Boswell', '1791'),
'taleoftwocities':('Charles Dickens', '1859'),
'narrativeofthecaptivity':('Mary Rowlandson', '1682'),
'ageofinnocence':('Edith Wharton', '1920'),
'soulsofblackfolk':('W. E. B. Du Bois', '1903'),
'ontheoriginofspecies':('Charles Darwin', '1859'),
'threecontributions':('Sigmund Freud', '1905'),
'waroftheworlds':('H. G. Wells','1898'),
'heartofdarkness':('Joseph Conrad', '1899'),
'metamorphosis':('Franz Kafka', '1915'),
'frankenstein':('Mary Shelley', '1818'),
'alicesadventures':('Lewis Carroll', '1865'),
'yellowwallpaper':('Charlotte Perkins Gilman', '1892'),
'autobiography':('John Stuart Mill', '1873'),
'anenquiry':('David Hume', '1748'),
'josephandrews':('Henry Fielding', '1742'),
'rasselas':('Samuel Johnson', '1759')
}

texts = {"middlemarch":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/middlemarch.txt"),
"portraitofalady":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/portraitofalady.txt"),
"huckleberryfinn":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/huckleberryfinn.txt"),
"emma":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/emma.txt"),
"mysteriesofudolpho":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/mysteriesofudolpho.txt"),
"uncletomscabin":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/uncletomscabin.txt"),
"lastofthemohicans":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/lastofthemohicans.txt"),
"mobydick":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/mobydick.txt"),
"janeeyre":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/mysteriesofudolpho.txt"),
"fallofthehouseofusher":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/janeeyre.txt"),
"nature":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/nature.txt"),
"sartorresartus":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/sartorresartus.txt"),
"walden":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/walden.txt"),
"adventuresofsherlockholmes":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/adventuresofsherlockholmes.txt"),
"shamela":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/shamela.txt"),
"pamela":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/pamela.txt"),
"gulliverstravels":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/gulliverstravels.txt"),
"callofthewild":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/callofthewild.txt"),
"treasureisland":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/treasureisland.txt"),
"scarletletter":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/scarletletter.txt"),
"olaudahequiano":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/olaudahequiano.txt"),
"frontierinamericanhistory":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/frontierinamericanhistory.txt"),
"lifeoffrederickdouglass":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/lifeoffrederickdouglass.txt"),
"annakarenina":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/annakarenina.txt"),
"donquixote":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/donquixote.txt"),
"brotherskaramazov":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/brotherskaramazov.txt"),
"lifeofjohnson":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/lifeofjohnson.txt"),
"taleoftwocities":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/taleoftwocities.txt"),
"narrativeofthecaptivity":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/narrativeofthecaptivity.txt"),
"ageofinnocence":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/ageofinnocence.txt"),
"soulsofblackfolk":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/soulsofblackfolk.txt"),
"ontheoriginofspecies":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/ontheoriginofspecies.txt"),
"threecontributions":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/threecontributions.txt"),
"waroftheworlds":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/waroftheworlds.txt"),
"heartofdarkness":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/heartofdarkness.txt"),
"metamorphosis":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/metamorphosis.txt"),
"frankenstein":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/frankenstein.txt"),
"alicesadventures":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/alicesadventures.txt"),
"yellowwallpaper":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/yellowwallpaper.txt"),
"autobiography":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/autobiography.txt"),
"anenquiry":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/anenquiry.txt"),
"josephandrews":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/josephandrews.txt"),
"rasselas":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/rasselas.txt"),

}


authorwords = {}

for title in title_keys:
	book = open(texts[title]).read().decode('UTF8')
	authorwords[title] = book


df = pd.DataFrame.from_dict(authorwords, orient = 'index', dtype = None)
df.columns = ['text']
text = df.text
lower_text = text.str.lower()


def remove_stopwords(content):
	cleaned = filter(lambda x: x not in ENGLISH_STOP_WORDS, content.split())
	return ' '.join(cleaned)
text = lower_text.apply(remove_stopwords)

def remove_punctuation(content):
	return filter(lambda x: x in string.ascii_letters+" ", content)
text = text.apply(remove_punctuation)

stemmer = SnowballStemmer('english')

def stem_words(content):
	stemmed_words = [stemmer.stem(word) for word in content.split()]
	return " ".join(stemmed_words)
text = text.apply(stem_words)

def lexdiverse(content):
	tokens = nltk.word_tokenize(content)
	lex_div = len(set(tokens))/len(tokens)
	return lex_div

lexical_diversity = text.apply(lexdiverse)





wb = Workbook()
ws1 = wb.active
ws1.cell(row=1, column=1).value = 'title'
ws1.cell(row=1, column=2).value = 'lexical_diversity'
ws1.cell(row=1, column=3).value = 'character length'

row_num = 2
index_num = 0
for x,y in zip(lexical_diversity, lower_text):
	ws1.cell(row = row_num, column = 1).value = str(lexical_diversity.index[index_num])
	ws1.cell(row = row_num, column = 2).value = x
	ws1.cell(row = row_num, column = 3).value = len(y)

	index_num += 1 
	row_num += 1

wb.save('ap_exceldata/ap_exceldata.xlsx')








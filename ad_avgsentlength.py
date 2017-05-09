from __future__ import division
import aa_authorwords as aa
import nltk
from openpyxl import Workbook
from openpyxl.compat import range


avg_sentlength = {}
for key in aa.title_keys:
	sent_token = nltk.sent_tokenize(aa.lc_authorwords[key])
	num_of_sent = len(sent_token)
	total_sentlength = 0
	for sent in sent_token:
		total_sentlength += len(sent)
	average_length = total_sentlength/num_of_sent
	avg_sentlength[key] = average_length

print avg_sentlength

wb = Workbook()
ws1 = wb.create_sheet('avg_sent_length')
ws1.cell(row=1, column=1).value = 'title'
ws1.cell(row=1, column=2).value = 'Avg Sent Length'

entry = 2
while entry <=(len(aa.title_keys)+1):
	for key in aa.title_keys:
		ws1.cell(row = entry, column = 1).value = key
		ws1.cell(row = entry, column = 2).value = avg_sentlength[key]
		entry +=1

wb.save('avg_sent_data.xlsx')
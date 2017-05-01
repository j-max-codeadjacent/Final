from __future__ import division
import start_end as se
import nltk
from openpyxl import Workbook
from openpyxl.compat import range
import lexicaldiversity as ld
import pandas as pd

#The title keys is a list created for easy access to the dicitonaries.
title_keys = [
#George Eliot
"adambede", 'themillonthefloss', 'silasmarner', 'romola', 'felixholt', 'middlemarch', "deronda", 
#Henry James
"roderickhudson", "theamericans", 'theeuropeans', 'confidence', 'washingtonsquare','portraitofaladyvol1', 'portraitofaladyvol2', 
'bostonians_v1','bostonians_v2','reverberator','tragicmuse','spoilsofpoynton','thegoldenbowl', 
#Mark Twain
"theinnocentsabroad", "roughingit", "theadventuresoftomsawyer", 'lifeonthemississippi', 'aconnecticutyankee', 
'thetragedyofpuddnhead', 'joanofarcvol1', 'joanofarcvol2'
]

titles = {
#George Eliot
"adambede":("George Eliot", "1859"), 
'themillonthefloss':("George Eliot", "1860"), 
'silasmarner':("George Eliot", "1861"), 
'romola':("George Eliot", "1863"), 
'felixholt':("George Eliot", "1866"), 
'middlemarch':("George Eliot", "1872"), 
"deronda":("George Eliot", "1876"), 
#Henry James
"roderickhudson":("Henry James", "1871"), 
"theamericans":("Henry James", "1877"), 
'theeuropeans':("Henry James", "1878"), 
'confidence':("Henry James", "1879"), 
'washingtonsquare':('Henry James', '1880'),
'portraitofaladyvol1':("Henry James", "1881"), 
'portraitofaladyvol2':("Henry James", "1881"), 
'bostonians_v1':('Henry James', '1886'),
'bostonians_v2':('Henry James', '1886'),
'reverberator':('Henry James', '1888'),
'tragicmuse':('Henry James', '1890'),
'spoilsofpoynton':('Henry James', '1898'),
'thegoldenbowl':("Henry James", "1904"), 
#Mark Twain
"theinnocentsabroad":("Mark Twain", '1869'), 
"roughingit":("Mark Twain", '1872'), 
"theadventuresoftomsawyer":("Mark Twain", '1876'), 
'lifeonthemississippi':("Mark Twain", '1883'), 
'aconnecticutyankee':("Mark Twain", '1889'), 
'thetragedyofpuddnhead':("Mark Twain", '1894'), 
'joanofarcvol1':("Mark Twain", '1896'), 
'joanofarcvol2':("Mark Twain", '1896')
}

texts = {"adambede":("/Users/jmax.barry/Documents/Coding/DataScience/Final/Texts/GeorgeEliot/adambede.txt"),
"themillonthefloss":("/Users/jmax.barry/Documents/Coding/DataScience/Final/Texts/GeorgeEliot/themillonthefloss.txt"),
"silasmarner":("/Users/jmax.barry/Documents/Coding/DataScience/Final/Texts/GeorgeEliot/silasmarner.txt"),
"romola":("/Users/jmax.barry/Documents/Coding/DataScience/Final/Texts/GeorgeEliot/romola.txt"),
"felixholt":("/Users/jmax.barry/Documents/Coding/DataScience/Final/Texts/GeorgeEliot/felixholt.txt"),
"middlemarch":("/Users/jmax.barry/Documents/Coding/DataScience/Final/Texts/GeorgeEliot/middlemarch.txt"),
"deronda": ("/Users/jmax.barry/Documents/Coding/DataScience/Final/Texts/GeorgeEliot/deronda.txt"),
#Henry James
"roderickhudson":("/Users/jmax.barry/Documents/Coding/DataScience/Final/Texts/HenryJames/roderickhudson.txt"),
"theamericans":("/Users/jmax.barry/Documents/Coding/DataScience/Final/Texts/HenryJames/theamericans.txt"),
"theeuropeans":("/Users/jmax.barry/Documents/Coding/DataScience/Final/Texts/HenryJames/theeuropeans.txt"),
"confidence":("/Users/jmax.barry/Documents/Coding/DataScience/Final/Texts/HenryJames/confidence.txt"),
"washingtonsquare":("/Users/jmax.barry/Documents/Coding/DataScience/Final/Texts/HenryJames/washingtonsquare.txt"),
"portraitofaladyvol1":("/Users/jmax.barry/Documents/Coding/DataScience/Final/Texts/HenryJames/portraitofaladyvol1.txt"),
"portraitofaladyvol2":("/Users/jmax.barry/Documents/Coding/DataScience/Final/Texts/HenryJames/portraitofaladyvol2.txt"),
"bostonians_v1":("/Users/jmax.barry/Documents/Coding/DataScience/Final/Texts/HenryJames/bostonians_v1.txt"),
"bostonians_v2":("/Users/jmax.barry/Documents/Coding/DataScience/Final/Texts/HenryJames/bostonians_v2.txt"),
"reverberator":("/Users/jmax.barry/Documents/Coding/DataScience/Final/Texts/HenryJames/reverberator.txt"),
"tragicmuse":("/Users/jmax.barry/Documents/Coding/DataScience/Final/Texts/HenryJames/tragicmuse.txt"),
"spoilsofpoynton":("/Users/jmax.barry/Documents/Coding/DataScience/Final/Texts/HenryJames/spoilsofpoynton.txt"),
"thegoldenbowl":("/Users/jmax.barry/Documents/Coding/DataScience/Final/Texts/HenryJames/thegoldenbowl.txt"),
#Mark Twain
"theinnocentsabroad":("/Users/jmax.barry/Documents/Coding/DataScience/Final/Texts/MarkTwain/theinnocentsabroad.txt"),
"roughingit":("/Users/jmax.barry/Documents/Coding/DataScience/Final/Texts/MarkTwain/roughingit.txt"),
"theadventuresoftomsawyer":("/Users/jmax.barry/Documents/Coding/DataScience/Final/Texts/MarkTwain/theadventuresoftomsawyer.txt"),
"lifeonthemississippi":("/Users/jmax.barry/Documents/Coding/DataScience/Final/Texts/MarkTwain/lifeonthemississippi.txt"),
"aconnecticutyankee":("/Users/jmax.barry/Documents/Coding/DataScience/Final/Texts/MarkTwain/aconnecticutyankee.txt"),
"thetragedyofpuddnhead":("/Users/jmax.barry/Documents/Coding/DataScience/Final/Texts/MarkTwain/thetragedyofpuddnhead.txt"),
"joanofarcvol1":("/Users/jmax.barry/Documents/Coding/DataScience/Final/Texts/MarkTwain/joanofarcvol1.txt"),
"joanofarcvol2":("/Users/jmax.barry/Documents/Coding/DataScience/Final/Texts/MarkTwain/joanofarcvol2.txt")
}

start_end = {'adambede': ('Book One', 'SELECTED BIBLIOGRAPHY'), 
'themillonthefloss': ('Book I', 'End of the Project Gutenberg EBook'), 
'silasmarner': ('PART ONE', 'End of the Project Gutenberg EBook'), 
'romola': ('PART ONE', 'End of the Project Gutenberg EBook'), 
'felixholt': ('INTRODUCTION', 'THE END'), 
'middlemarch': ('PRELUDE', 'End of the Project Gutenberg EBook'), 
'deronda':('BOOK I', 'End of the Project Gutenberg'),
'roderickhudson':('CHAPTER I', 'End of the Project Gutenberg EBook'),
'theamericans':('On a brilliant', 'End of the Project Gutenberg EBook'),
'theeuropeans':('A narrow grave-yard', 'The End'),
'confidence':('CHAPTER I', 'End of the Project Gutenberg EBook'),
'washingtonsquare':('DURING a portion', '***END OF THE PROJECT'),
'portraitofaladyvol1':('Under certain circumstances', 'End of the Project Gutenberg EBook'),
'portraitofaladyvol2':('CHAPTER XXVIII', 'End of the Project Gutenberg EBook'),
'bostonians_v1':('Olive will come down', 'END OF VOL. I'),
'bostonians_v2':('A little more than an', 'THE END'),
'reverberator':('I guess my daughter', 'End of the Project Gutenberg'),
'tragicmuse':('I profess a certain', '***END OF THE PROJECT'),
'spoilsofpoynton':('Mrs. Gereth had said she', "Henry James's Books."),
'thegoldenbowl':('BOOK FIRST: THE PRINCE', 'End of the Project Gutenberg EBook'),
'theinnocentsabroad':('CHAPTER I', 'End of the Project Gutenberg EBook'),
'roughingit':('My brother had just been', 'End of the Project Gutenberg EBook'),
'theadventuresoftomsawyer':('TOM!', 'End of the Project Gutenberg EBook'),
'lifeonthemississippi':('The River and Its History', 'End of the Project Gutenberg EBook'),
'aconnecticutyankee':('The ungentle laws', 'End of the Project Gutenberg EBook'),
'thetragedyofpuddnhead':('A Whisper', 'Transcriber\'s Notes'),
'joanofarcvol1':('By The Sieur Louis De Conte', 'End of the Project Gutenberg EBook'),
'joanofarcvol2':('BOOK II -- IN COURT AND CAMP (Continued)', 'End of the Project Gutenberg EBook')
}


#book = raw_input("What book would you like to analyze?")
#text_file = texts[book]



#each entry in authorwords consists of the PG book minus the meta-text.
authorwords = {}
for x in titles:
	book = texts[x]
	raw_book = open(book, 'r')
	start_read = start_end[str(x)][0] 
	end_read = start_end[str(x)][1]
	book_authorwords = se.rawifier(raw_book, start_read, end_read)
	book_authorwords = book_authorwords.decode('utf8')
	authorwords[x]=book_authorwords


def avgsentlength():
	avsentlength_forXL = {}
	#title_count = 0
	for key in title_keys:
		sent_tokens = nltk.sent_tokenize(authorwords[key])
		total_length = 0
		for sentence in sent_tokens:
			total_length += len(sentence)
		avg_sent_length = total_length/len(sent_tokens)
		avsentlength_forXL[key] = avg_sent_length
	return avsentlength_forXL
		#print "%r has an average sentence length of %r" %(book, avg_sent_length)
#avsentlength_forXL = avgsentlength()
#print avsentlength_forXL

def lexicaldiversity():
	lexdiv_forXL = {}
	for key in title_keys:
		book = authorwords[key]
		tokens = nltk.word_tokenize(book)
		lex_div = len(tokens)/len(book)
 		lexdiv_forXL[key] =  lex_div
 	return lexdiv_forXL
#lexdiv_forXL = lexicaldiversity()
#print lexdiv_forXL	

wb = Workbook()
ws1 = wb.active
ws1.title = 'initialdata'
ws1.sheet_properties.tabColor = "ffa500"

ws1.cell(row=1, column=1).value = 'author'
ws1.cell(row=1, column=2).value = 'title'
ws1.cell(row=1, column=3).value = 'pub_date'
ws1.cell(row=1, column=4).value = 'avg_sent_length'
ws1.cell(row=1, column=5).value = 'lexical_diversity'

#creates the row headings
entry = 2
while entry <= (len(titles)+1): #thought I needed to add +1 to account for starting at entry 2
	for key in title_keys:
		ws1.cell(row=entry, column=1).value = titles[key][0]
		ws1.cell(row=entry, column=2).value = key
		ws1.cell(row=entry, column=3).value = titles[key][1]
		entry += 1

#creates the avsentlength_forXL column
def create_XL():	
	entry = 2
	while entry <= (len(title_keys)+1): #thought I needed to add +1 to account for starting at entry 2
		for key in title_keys:
			ws1.cell(row=entry, column=4).value = avsentlength_forXL[key]
			entry += 1

	entry = 2
	while entry <= (len(title_keys)+1): #thought I needed to add +1 to account for starting at entry 2
		for key in title_keys:
			ws1.cell(row=entry, column=5).value = lexdiv_forXL[key]
			entry += 1

	entry = 2
	while entry <= (len(title_keys)+1): #thought I needed to add +1 to account for starting at entry 2
		for key in title_keys:
			ws1.cell(row=entry, column=6).value = "houndstooth"
			entry += 1

	wb.save('novel_data.xlsx')


#testing creating data frame objects for each book, starting with Adam Bede.
def sentences_forDF():
	sent_tokens = nltk.sent_tokenize(authorwords[title_keys[0]])
	print type(sent_tokens)
	return sent_tokens	
sentences_forDF = sentences_forDF()

def lowercasesentences_forDF():
	lc_sentences = []
	for sent in sentences_forDF:
		lc_sent = sent.lower()
		lc_sentences.append(lc_sent)
	print lc_sentences[2]	
	return lc_sentences
lowercasesentences_forDF = lowercasesentences_forDF()
#Average sentence length includes white spaces and punctuation, including quotation marks.
def sentencelengths_forDF():
	sentence_lengths = []
	for x in sentences_forDF:
		sentence_lengths.append(len(x)) 			
 	return sentence_lengths
sentencelengths_forDF = sentencelengths_forDF()


def lexdiversity_forDF():
	tokens = nltk.word_tokenize(title_keys[0])
	lexdiv_forDF = len(tokens)/len(book)
 	return lexdiv_forDF
lexdiversity_forDF = lexdiversity_forDF()

df = pd.DataFrame.from_records({"Sentences": sentences_forDF,
 "Lower Case Sentences": lowercasesentences_forDF,
 "Sentence Lengths": sentencelengths_forDF,
 })
print df.head()
"""labels = ("Sentence Length")
df = pd.DataFrame(lexdiversity_forDF, 
	avgsentlength_forDF, 
	sentences_forDF, 
	lowercasesentences_forDF, 
	)
print df.head()
print df.tail()
"""
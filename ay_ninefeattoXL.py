import nltk
from openpyxl import Workbook
import regex as re

title_keys = [
'middlemarch', 
"portraitofalady", 
"huckleberryfinn",
'emma',
'mysteriesofudolpho',
'uncletomscabin',
'lastofthemohicans',
'mobydick',
#'janeeyre', #this is not utf8
#'fallofthehouseofusher',
'nature',
'sartorresartus',
'walden',
'adventuresofsherlockholmes',
#'shamela',
'pamela',
'gulliverstravels',
'callofthewild',
'treasureisland',
'scarletletter',
'olaudahequiano',
'frontierinamericanhistory',
'lifeoffrederickdouglass',
'annakarenina',
'donquixote',
'brotherskaramazov',
'lifeofjohnson',
'taleoftwocities',
'narrativeofthecaptivity',
'ageofinnocence',
'soulsofblackfolk',
'ontheoriginofspecies',
'threecontributions',
'waroftheworlds',
'heartofdarkness',
'metamorphosis',
'frankenstein', 
'alicesadventures',
'yellowwallpaper',
'autobiography',
'anenquiry',
'josephandrews',
'rasselas'


]

titles = {
'middlemarch':("George Eliot", "1872"), 
'portraitofalady':("Henry James", "1881"),
'huckleberryfinn':("Mark Twain", '1884'),
'emma':('Jane Austen', '1815'),
'mysteriesofudolpho':('Ann Radcliffe', '1794'),
'uncletomscabin':('Harriet Beecher Stowe', '1852'),
'lastofthemohicans':('James Fenimore Cooper', '1826'),
'mobydick':('Herman Melville', '1851'),
'janeeyre':('Charlotte Bronte', '1847'),
'fallofthehouseofusher':('Edgar Allan Poe', '1839'),
'nature':('Ralph Waldo Emerson', '1836'),
'sartorresartus':('Thomas Carlyle', '1836'),
'walden':('Henry David Thoreau', '1854'),
'adventuresofsherlockholmes':('Arthur Conan Doyle', '1892'),
'shamela':('Henry Fielding', '1741'),
'pamela':('Samuel Richardson', '1740'),
'gulliverstravels':('Jonathan Swift', '1726'),
'callofthewild':('Jack London', '1903'),
'treasureisland':('Robert Louis Stevenson', '1883'),
'scarletletter':('Nathaniel Hawthorne', '1850'),
'olaudahequiano':('Olaudah Equiano','1789'),
'frontierinamericanhistory':('Frederick Jackson Turner', '1921'),
'lifeoffrederickdouglass':('Frederick Douglass', '1845'),
'annakarenina':('Leo Tolstoy', '1877'),
'donquixote':('Miguel de Cervantes', '1601'),
'brotherskaramazov':('Fyodor Dostoyevsky', '1880'),
'lifeofjohnson':('James Boswell', '1791'),
'taleoftwocities':('Charles Dickens', '1859'),
'narrativeofthecaptivity':('Mary Rowlandson', '1682'),
'ageofinnocence':('Edith Wharton', '1920'),
'soulsofblackfolk':('W. E. B. Du Bois', '1903'),
'ontheoriginofspecies':('Charles Darwin', '1859'),
'threecontributions':('Sigmund Freud', '1905'),
'waroftheworlds':('H. G. Wells','1898'),
'heartofdarkness':('Joseph Conrad', '1899'),
'metamorphosis':('Franz Kafka', '1915'),
'frankenstein':('Mary Shelley', '1818'),
'alicesadventures':('Lewis Carroll', '1865'),
'yellowwallpaper':('Charlotte Perkins Gilman', '1892'),
'autobiography':('John Stuart Mill', '1873'),
'anenquiry':('David Hume', '1748'),
'josephandrews':('Henry Fielding', '1742'),
'rasselas':('Samuel Johnson', '1759')
}

texts = {"middlemarch":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/middlemarch.txt"),
"portraitofalady":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/portraitofalady.txt"),
"huckleberryfinn":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/huckleberryfinn.txt"),
"emma":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/emma.txt"),
"mysteriesofudolpho":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/mysteriesofudolpho.txt"),
"uncletomscabin":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/uncletomscabin.txt"),
"lastofthemohicans":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/lastofthemohicans.txt"),
"mobydick":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/mobydick.txt"),
"janeeyre":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/mysteriesofudolpho.txt"),
"fallofthehouseofusher":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/janeeyre.txt"),
"nature":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/nature.txt"),
"sartorresartus":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/sartorresartus.txt"),
"walden":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/walden.txt"),
"adventuresofsherlockholmes":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/adventuresofsherlockholmes.txt"),
"shamela":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/shamela.txt"),
"pamela":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/pamela.txt"),
"gulliverstravels":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/gulliverstravels.txt"),
"callofthewild":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/callofthewild.txt"),
"treasureisland":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/treasureisland.txt"),
"scarletletter":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/scarletletter.txt"),
"olaudahequiano":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/olaudahequiano.txt"),
"frontierinamericanhistory":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/frontierinamericanhistory.txt"),
"lifeoffrederickdouglass":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/lifeoffrederickdouglass.txt"),
"annakarenina":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/annakarenina.txt"),
"donquixote":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/donquixote.txt"),
"brotherskaramazov":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/brotherskaramazov.txt"),
"lifeofjohnson":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/lifeofjohnson.txt"),
"taleoftwocities":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/taleoftwocities.txt"),
"narrativeofthecaptivity":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/narrativeofthecaptivity.txt"),
"ageofinnocence":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/ageofinnocence.txt"),
"soulsofblackfolk":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/soulsofblackfolk.txt"),
"ontheoriginofspecies":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/ontheoriginofspecies.txt"),
"threecontributions":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/threecontributions.txt"),
"waroftheworlds":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/waroftheworlds.txt"),
"heartofdarkness":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/heartofdarkness.txt"),
"metamorphosis":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/metamorphosis.txt"),
"frankenstein":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/frankenstein.txt"),
"alicesadventures":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/alicesadventures.txt"),
"yellowwallpaper":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/yellowwallpaper.txt"),
"autobiography":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/autobiography.txt"),
"anenquiry":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/anenquiry.txt"),
"josephandrews":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/josephandrews.txt"),
"rasselas":("/Users/jmax.barry/Documents/Coding/DataScience/Final/new_noquotes/rasselas.txt"),

}

authorwords = {}

for title in title_keys:
	book = open(texts[title]).read().decode('UTF8')
	author = titles[title][0]
	authorwords[title] = [author,book]

sent_byauthor = []
for x in authorwords:
	sent_tokens = nltk.sent_tokenize(authorwords[x][1])
	for sent in sent_tokens:
		
		avg_word_len = 0
		nopunc_sent = re.sub(ur"\p{P}+", "", sent) #removes unicode punctuation
		words = nltk.word_tokenize(sent)
		if len(words) == 0:
			avg_word_length = 'delete'
		else:			
			sent_lenby_character = 0
			for y in words:
				sent_lenby_character += len(y)
			if sent_lenby_character == 1:
				avg_word_len = 'delete'
			else:
				avg_word_len = float(sent_lenby_character)/len(words)
		

		sent_length = len(sent)
		question = 0
		ex_point = 0
		colon = 0
		semi_colon = 0
		if '?' in sent:
			question = 1
		if '!' in sent:
			ex_point = 1
		if ':' in sent:
			colon = 1
		if ';' in sent:
			semi_colon = 1

		if avg_word_len != 'delete':		
			sent_byauthor.append([x, authorwords[x][0], sent, sent_length, avg_word_len, 
									question, ex_point, colon, semi_colon])
		
		

wb = Workbook()
ws1 = wb.active
ws1.title = 'Features'
ws1.cell(row=1, column=1).value = 'title'
ws1.cell(row=1, column=2).value = 'sentence'
ws1.cell(row=1, column=3).value = 'length'
ws1.cell(row=1, column=4).value = 'avg_word_length'
ws1.cell(row=1, column=5).value = 'q_mark'
ws1.cell(row=1, column=6).value = 'ex_point'
ws1.cell(row=1, column=7).value = 'colon'
ws1.cell(row=1, column=8).value = 'semi_colon'



row_num = 2
for x in sent_byauthor:
	ws1.cell(row=row_num, column=1).value = x[1]
	ws1.cell(row=row_num, column=2).value = x[2]
	ws1.cell(row=row_num, column=3).value = x[3]
	ws1.cell(row=row_num, column=4).value = x[4]
	ws1.cell(row=row_num, column=5).value = x[5]
	ws1.cell(row=row_num, column=6).value = x[6]
	ws1.cell(row=row_num, column=7).value = x[7]
	ws1.cell(row=row_num, column=8).value = x[8]
	row_num += 1

wb.save('exceldata/sentenceData/ay.xlsx')




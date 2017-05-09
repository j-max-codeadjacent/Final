# -- coding: utf-8 
from __future__ import division
import start_end as se
import nltk
from openpyxl import Workbook
from openpyxl.compat import range
import pandas as pd

title_keys = [
'middlemarch', 
"portraitofalady", 
"huckleberryfinn",
'emma',
'mysteriesofudolpho',
'uncletomscabin',
'lastofthemohicans',
'mobydick',
'janeeyre',
'fallofthehouseofusher',
'nature',
'sartorresartus',
'walden',
'adventuresofsherlockholmes',
'shamela',
'pamela',
'gulliverstravels',
'callofthewild',
'treasureisland',
'scarletletter',
'olaudahequiano',
'frontierinamericanhistory',
'lifeoffrederickdouglass',
'annakarenina',
'donquixote',
'brotherskaramazov',
'lifeofjohnson',
'taleoftwocities',
'narrativeofthecaptivity',
'ageofinnocence',
'soulsofblackfolk',
'ontheoriginofspecies',
'threecontributions',
'waroftheworlds',
'heartofdarkness',
'metamorphosis',
'frankenstein', 
'alicesadventures',
'yellowwallpaper',
'autobiography',
'anenquiry',
'josephandrews',
'rasselas'


]

titles = {
'middlemarch':("George Eliot", "1872"), 
'portraitofalady':("Henry James", "1881"),
'huckleberryfinn':("Mark Twain", '1884'),
'emma':('Jane Austen', '1815'),
'mysteriesofudolpho':('Ann Radcliffe', '1794'),
'uncletomscabin':('Harriet Beecher Stowe', '1852'),
'lastofthemohicans':('James Fenimore Cooper', '1826'),
'mobydick':('Herman Melville', '1851'),
'janeeyre':('Charlotte Brontë', '1847'),
'fallofthehouseofusher':('Edgar Allan Poe', '1839'),
'nature':('Ralph Waldo Emerson', '1836'),
'sartorresartus':('Thomas Carlyle', '1836'),
'walden':('Henry David Thoreau', '1854'),
'adventuresofsherlockholmes':('Arthur Conan Doyle', '1892'),
'shamela':('Henry Fielding', '1741'),
'pamela':('Samuel Richardson', '1740'),
'gulliverstravels':('Jonathan Swift', '1726'),
'callofthewild':('Jack London', '1903'),
'treasureisland':('Robert Louis Stevenson', '1883'),
'scarletletter':('Nathaniel Hawthorne', '1850'),
'olaudahequiano':('Olaudah Equiano','1789'),
'frontierinamericanhistory':('Frederick Jackson Turner', '1921'),
'lifeoffrederickdouglass':('Frederick Douglass', '1845'),
'annakarenina':('Leo Tolstoy', '1877'),
'donquixote':('Miguel de Cervantes', '1601'),
'brotherskaramazov':('Fyodor Dostoyevsky', '1880'),
'lifeofjohnson':('James Boswell', '1791'),
'taleoftwocities':('Charles Dickens', '1859'),
'narrativeofthecaptivity':('Mary Rowlandson', '1682'),
'ageofinnocence':('Edith Wharton', '1920'),
'soulsofblackfolk':('W. E. B. Du Bois', '1903'),
'ontheoriginofspecies':('Charles Darwin', '1859'),
'threecontributions':('Sigmund Freud', '1905'),
'waroftheworlds':('H. G. Wells','1898'),
'heartofdarkness':('Joseph Conrad', '1899'),
'metamorphosis':('Franz Kafka', '1915'),
'frankenstein':('Mary Shelley', '1818'),
'alicesadventures':('Lewis Carroll', '1865'),
'yellowwallpaper':('Charlotte Perkins Gilman', '1892'),
'autobiography':('John Stuart Mill', '1873'),
'anenquiry':('David Hume', '1748'),
'josephandrews':('Henry Fielding', '1742'),
'rasselas':('Samuel Johnson', '1759')
}

texts = {"middlemarch":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/middlemarch.txt"),
"portraitofalady":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/portraitofalady.txt"),
"huckleberryfinn":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/huckleberryfinn.txt"),
"emma":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/emma.txt"),
"mysteriesofudolpho":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/mysteriesofudolpho.txt"),
"uncletomscabin":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/uncletomscabin.txt"),
"lastofthemohicans":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/lastofthemohicans.txt"),
"mobydick":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/mobydick.txt"),
"janeeyre":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/mysteriesofudolpho.txt"),
"fallofthehouseofusher":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/janeeyre.txt"),
"nature":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/nature.txt"),
"sartorresartus":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/sartorresartus.txt"),
"walden":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/walden.txt"),
"adventuresofsherlockholmes":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/adventuresofsherlockholmes.txt"),
"shamela":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/shamela.txt"),
"pamela":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/pamela.txt"),
"gulliverstravels":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/gulliverstravels.txt"),
"callofthewild":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/callofthewild.txt"),
"treasureisland":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/treasureisland.txt"),
"scarletletter":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/scarletletter.txt"),
"olaudahequiano":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/olaudahequiano.txt"),
"frontierinamericanhistory":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/frontierinamericanhistory.txt"),
"lifeoffrederickdouglass":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/lifeoffrederickdouglass.txt"),
"annakarenina":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/annakarenina.txt"),
"donquixote":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/donquixote.txt"),
"brotherskaramazov":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/brotherskaramazov.txt"),
"lifeofjohnson":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/lifeofjohnson.txt"),
"taleoftwocities":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/taleoftwocities.txt"),
"narrativeofthecaptivity":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/narrativeofthecaptivity.txt"),
"ageofinnocence":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/ageofinnocence.txt"),
"soulsofblackfolk":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/soulsofblackfolk.txt"),
"ontheoriginofspecies":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/ontheoriginofspecies.txt"),
"threecontributions":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/threecontributions.txt"),
"waroftheworlds":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/waroftheworlds.txt"),
"heartofdarkness":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/heartofdarkness.txt"),
"metamorphosis":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/metamorphosis.txt"),
"frankenstein":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/frankenstein.txt"),
"alicesadventures":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/alicesadventures.txt"),
"yellowwallpaper":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/yellowwallpaper.txt"),
"autobiography":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/autobiography.txt"),
"anenquiry":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/anenquiry.txt"),
"josephandrews":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/josephandrews.txt"),
"rasselas":("/Users/jmax.barry/Documents/Coding/DataScience/Final/noquotes_texts/rasselas.txt"),

}

start_end = {
'middlemarch': ('PRELUDE', 'End of the Project Gutenberg EBook'), 
'portraitofalady':('Under certain circumstances', 'End of the Project Gutenberg EBook'),
'huckleberryfinn':('know about me without', 'End of Project Gutenberg'),
'emma':('Emma Woodhouse, handsome, clever,','End of the Project Gutenberg EBook of Emma'),
'mysteriesofudolpho':('On the pleasant banks of the Garonne', "End of Project Gutenberg's The Mysteries"),
'uncletomscabin':('Late in the afternoon of a chilly day', 'End of Project Gutenberg'),
'lastofthemohicans':('It is believed that the scene', 'End of Project Gutenberg'),
'mobydick':('The pale Usher—threadbare in coat', 'End of Project Gutenberg'),
'janeeyre':('There was no possibility of taking a walk', '***END OF THE PROJECT GUTENBERG'),
'fallofthehouseofusher':('During the whole of a dull', '* Watson, Dr Percival, Spallanzani,'),
'nature':('OUR age is retrospective. It builds', 'End of the Project Gutenberg'),
'sartorresartus':('Considering our present advanced state', 'existed together, though in a state of quarrel?'),
'walden':('When I wrote the following pages', 'End of the Project Gutenberg'),
'adventuresofsherlockholmes':('To Sherlock Holmes she is always THE woman.', 'End of the Project Gutenberg'),
'shamela':('In which, the many notorious FALSHOODS and MISREPRSENTATIONS of a Book called', 'End of the Project Gutenberg'),
'pamela':('DEAR FATHER AND MOTHER,', 'End of Project Gutenberg'),
'gulliverstravels':('I hope you will be ready to own publicly', 'footnotes'),
'callofthewild':('Buck did not read the newspapers', 'End of the Project Gutenberg'),
'treasureisland':('Squire Trelawney, Doctor Livesey,', 'End of Project Gutenberg'),
'scarletletter':('It is a little remarkable, that', 'End of Project Gutenberg'),
'olaudahequiano':('I believe it is difficult for those','End of the Project Gutenberg EBook'),
'frontierinamericanhistory':('In a recent bulletin of the Superintendent of the Census', 'lest that freedom be lost forever'),
'lifeoffrederickdouglass':('In the month of August, 1841', 'FREDERICK DOUGLASS. LYNN, _Mass., April_ 28, 1845'),
'annakarenina':('Happy families are all alike', '*** END OF THIS PROJECT GUTENBERG'),
'donquixote':('Idle reader: thou mayest believe me without', 'End of the Project Gutenberg EBook'),
'brotherskaramazov':('Alexey Fyodorovitch Karamazov was the third son of Fyodor', 'boys took up his'),
'lifeofjohnson':('Phillips Brooks once told the boys', 'End of the Project Gutenberg EBook'),
'taleoftwocities':('It was the best of times,', 'End of the Project Gutenberg EBook '),
'narrativeofthecaptivity':('The sovereignty and goodness of GOD,', 'EEnd of Project Gutenberg'),
'ageofinnocence':('On a January evening of the early seventies', 'up slowly and walked back alone to his hotel'),
'soulsofblackfolk':('Herein lie buried many things', 'End of Project Gutenberg'),
'ontheoriginofspecies':('When on board H.M.S. ', 'and most wonderful have been, and are being, evolved.'),
'threecontributions':('The fact of sexual need', 'normal or the pathological..'),
'waroftheworlds':('No one would have believed','End of the Project Gutenberg'),
'heartofdarkness':('The Nellie, a cruising yawl,', 'End of the Project Gutenberg'),
'metamorphosis':('One morning, when Gregor Samsa', 'End of the Project Gutenberg'),
'frankenstein':('You will rejoice to hear that', 'End of the Project Gutenberg'),
'alicesadventures':('Alice was beginning to get', 'End of Project Gutenberg'),
'yellowwallpaper':('It is very seldom that mere', 'End of Project Gutenberg'),
'autobiography':('It seems proper that I should', 'for the present, this memoir may close.'),
'anenquiry':('1. Moral philosophy, or the science', 'contain nothing but sophistry'),
'josephandrews':('Of writing lives in general', 'to make his appearance in high life.'),
'rasselas':('YE who listen with credulity', 'to return to Abyssinia.')
}

#book = raw_input("What book would you like to analyze?")
#text_file = texts[book]



#each entry in authorwords consists of the PG book minus the meta-text.
authorwords = {}
for x in titles:
	book = texts[x]
	raw_book = open(book, 'r')
	start_read = start_end[str(x)][0] 
	end_read = start_end[str(x)][1]
	book_authorwords = se.rawifier(raw_book, start_read, end_read)
	book_authorwords = book_authorwords.decode('utf8')
	authorwords[x]=book_authorwords
print authorwords['huckleberryfinn']



def avgsentlength():
	avsentlength_forXL = {}
	#title_count = 0
	for key in title_keys:
		sent_tokens = nltk.sent_tokenize(authorwords[key])
		total_length = 0
		for sentence in sent_tokens:
			total_length += len(sentence)
		avg_sent_length = total_length/len(sent_tokens)
		avsentlength_forXL[key] = avg_sent_length
	return avsentlength_forXL
		#print "%r has an average sentence length of %r" %(book, avg_sent_length)
avsentlength_forXL = avgsentlength()
#print avsentlength_forXL


def lexicaldiversity():
	lexdiv_forXL = {}
	for key in title_keys:
		book = authorwords[key]
		tokens = nltk.word_tokenize(book)
		lex_div = len(tokens)/len(book)
 		lexdiv_forXL[key] =  lex_div
 	return lexdiv_forXL
lexdiv_forXL = lexicaldiversity()
#print lexdiv_forXL	


wb = Workbook()
ws1 = wb.active
ws1.title = 'initialdata'
ws1.sheet_properties.tabColor = "ffa500"

ws1.cell(row=1, column=1).value = 'author'
ws1.cell(row=1, column=2).value = 'title'
ws1.cell(row=1, column=3).value = 'pub_date'
ws1.cell(row=1, column=4).value = 'avg_sent_length'
ws1.cell(row=1, column=5).value = 'lexical_diversity'

#creates the row headings
entry = 2
while entry <= (len(titles)+1): #thought I needed to add +1 to account for starting at entry 2
	for key in title_keys:
		ws1.cell(row=entry, column=1).value = titles[key][0]
		ws1.cell(row=entry, column=2).value = key
		ws1.cell(row=entry, column=3).value = titles[key][1]
		entry += 1

#creates the avsentlength_forXL column

entry = 2
while entry <= (len(title_keys)+1): #thought I needed to add +1 to account for starting at entry 2
	for key in title_keys:
		ws1.cell(row=entry, column=4).value = avsentlength_forXL[key]
		entry += 1

entry = 2
while entry <= (len(title_keys)+1): #thought I needed to add +1 to account for starting at entry 2
	for key in title_keys:
		ws1.cell(row=entry, column=5).value = lexdiv_forXL[key]
		entry += 1

entry = 2
while entry <= (len(title_keys)+1): #thought I needed to add +1 to account for starting at entry 2
	for key in title_keys:
		ws1.cell(row=entry, column=6).value = "Elvis"
		entry += 1

wb.save('novel_data.xlsx')



#testing creating data frame objects for each book, starting with Adam Bede.
def sentences_forDF(title):
	sent_tokens = nltk.sent_tokenize(authorwords[title])
	print type(sent_tokens)
	return sent_tokens	
#title = raw_input("What book would you like to analyze?: %r" %title_keys) 
#sentences_forDF = sentences_forDF(title)

def lowercasesentences_forDF():
	lc_sentences = []
	for sent in sentences_forDF:
		lc_sent = sent.lower()
		lc_sentences.append(lc_sent)
	print lc_sentences[2]	
	return lc_sentences
#lowercasesentences_forDF = lowercasesentences_forDF()
#Average sentence length includes white spaces and punctuation, including quotation marks.
def sentencelengths_forDF():
	sentence_lengths = []
	for x in sentences_forDF:
		sentence_lengths.append(len(x)) 			
 	return sentence_lengths
#sentencelengths_forDF = sentencelengths_forDF()


def lexdiversity_forDF(title):
	tokens = nltk.word_tokenize(title)
	lexdiv_forDF = len(tokens)/len(book)
 	return lexdiv_forDF


"""df = pd.DataFrame.from_records({"Sentences": sentences_forDF,
 "Lower Case Sentences": lowercasesentences_forDF,
 "Sentence Lengths": sentencelengths_forDF,
 })
print df.head()
labels = ("Sentence Length")
df = pd.DataFrame(lexdiversity_forDF, 
	avgsentlength_forDF, 
	sentences_forDF, 
	lowercasesentences_forDF, 
	)
print df.head()
print df.tail()
"""